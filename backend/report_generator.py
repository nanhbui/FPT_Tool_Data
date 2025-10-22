"""
Module tạo báo cáo Excel với định dạng và dynamic linking
PHIÊN BẢN CẬP NHẬT: Hỗ trợ multi-file mode với cột MONTH
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from config import COLORS, NUMBER_FORMAT, INTEGER_FORMAT


class ReportGenerator:
    """Class tạo báo cáo Excel"""

    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.current_row = 1
        self.project_code_row_map = {}
        self.ratecard_col_letter = None

    def create_workbook(self):
        """Tạo workbook mới"""
        self.workbook = Workbook()
        # Xóa sheet mặc định để tránh thừa
        if self.workbook.active is not None:
            self.workbook.remove(self.workbook.active)

    def get_month_name(self, month):
        """Chuyển số tháng thành tên viết tắt"""
        months = [
            "Jan",
            "Feb",
            "Mar",
            "Apr",
            "May",
            "Jun",
            "Jul",
            "Aug",
            "Sep",
            "Oct",
            "Nov",
            "Dec",
        ]
        return months[month - 1]

    def create_project_code_sheet(self, df_project_code, all_project_codes):
        """Tạo sheet Project_Code từ DataFrame và đảm bảo có đủ tất cả Project Codes"""
        if self.workbook is None:
            return

        import pandas as pd

        # Lấy danh sách Project Code hiện có trong df (chuẩn hóa)
        existing_codes = set(
            df_project_code["Project Code"].astype(str).str.strip().tolist()
        )

        # Tìm các code thiếu (xuất hiện ở Project Report nhưng không có trong Project_Code)
        missing_codes = set(all_project_codes) - existing_codes

        # Nếu có code thiếu, thêm vào DataFrame với Ratecard = 0
        if missing_codes:
            print(
                f"  ⚠ Thêm {len(missing_codes)} Project Codes thiếu vào sheet Project_Code với Ratecard = 0:"
            )
            for code in sorted(missing_codes):
                print(f"    - {code}")

            missing_rows = []
            for code in missing_codes:
                row_dict = {"Project Code": code, "Ratecard": 0}

                # Thêm các columns khác nếu có trong df_project_code (để trống)
                for col in df_project_code.columns:
                    if col not in ["Project Code", "Ratecard"]:
                        row_dict[col] = ""

                missing_rows.append(row_dict)

            df_missing = pd.DataFrame(missing_rows)
            df_project_code = pd.concat(
                [df_project_code, df_missing], ignore_index=True
            )
            df_project_code = df_project_code.sort_values("Project Code").reset_index(
                drop=True
            )

        ws = self.workbook.create_sheet(title="Project_Code", index=0)

        ratecard_col_idx = None
        project_code_col_idx = None

        for idx, col_name in enumerate(df_project_code.columns, start=1):
            if col_name.strip() == "Ratecard":
                ratecard_col_idx = idx
                self.ratecard_col_letter = get_column_letter(idx)
            if col_name.strip() == "Project Code":
                project_code_col_idx = idx

        if ratecard_col_idx is None:
            self.ratecard_col_letter = "B"

        header_fill = PatternFill(
            start_color=COLORS["fixed_header"],
            end_color=COLORS["fixed_header"],
            fill_type="solid",
        )
        header_font = Font(bold=True, size=11, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, column_name in enumerate(df_project_code.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            if not isinstance(cell, MergedCell):
                cell.value = column_name
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

        self.project_code_row_map = {}

        for row_idx, row_data in enumerate(
            df_project_code.itertuples(index=False), start=2
        ):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
                    cell.border = thin_border

                    column_name = df_project_code.columns[col_idx - 1]
                    if column_name == "Ratecard":
                        cell.number_format = NUMBER_FORMAT

            project_code = str(
                ws.cell(row=row_idx, column=project_code_col_idx).value
            ).strip()
            self.project_code_row_map[project_code] = row_idx

        for col_idx, column_name in enumerate(df_project_code.columns, start=1):
            col_letter = get_column_letter(col_idx)
            if column_name == "Project Code":
                ws.column_dimensions[col_letter].width = 25
            elif column_name == "Ratecard":
                ws.column_dimensions[col_letter].width = 15
            else:
                ws.column_dimensions[col_letter].width = 20

    def get_revenue_formula(self, project_code):
        """Tạo Excel formula để reference đến Ratecard"""
        if project_code in self.project_code_row_map and self.ratecard_col_letter:
            row_num = self.project_code_row_map[project_code]
            return f"=Project_Code!${self.ratecard_col_letter}${row_num}"
        else:
            return 0

    def generate_report_two_sheets(
        self, df_input, df_monthly, month_list, output_path, df_project_code=None
    ):
        """
        Tạo báo cáo 2 sheets:
        1. Project Report: records gốc
        2. Summary: Metrics theo tháng (allocate)
        """
        self.create_workbook()

        if self.workbook is None:
            print("✗ Không thể tạo workbook")
            return

        # Lấy tất cả Project Codes từ df_input
        all_project_codes = (
            df_input["Project Code"].astype(str).str.strip().unique().tolist()
        )

        # 1. Tạo sheet Project_Code (với tất cả project codes)
        if df_project_code is not None:
            self.create_project_code_sheet(df_project_code, all_project_codes)

        # 2. Tạo sheet Project Report (records gốc)
        print("  Tạo sheet Project Report...")
        self._create_project_report_sheet(df_input, df_project_code)

        # 3. Tạo sheet Summary (metrics theo tháng)
        print("  Tạo sheet Summary...")
        self._create_summary_sheet(df_monthly, month_list)

        # 4. Lưu file
        if self.workbook is not None:
            self.workbook.save(output_path)
            print(f"✓ Báo cáo đã được tạo: {output_path}")
        else:
            print("✗ Không thể lưu file")

    def _create_project_report_sheet(self, df_input, df_project_code):
        """Tạo sheet Project Report với records gốc và bảng Total đẹp"""

        if self.workbook is None:
            return

        ws = self.workbook.create_sheet(title="Project Report", index=1)

        header_fill = PatternFill(
            start_color=COLORS["fixed_header"],
            end_color=COLORS["fixed_header"],
            fill_type="solid",
        )
        header_font = Font(bold=True, size=12, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        thick_border = Border(
            left=Side(style="medium"),
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium"),
        )

        # Kiểm tra xem có cột Month_Label không (multi-file mode)
        has_month_label = "Month_Label" in df_input.columns

        if has_month_label:
            headers = [
                "NO",
                "MONTH",
                "ACCOUNT",
                "MAIL",
                "PROJECT CODE",
                "AI PROJECT",
                "REVENUE",
                "CALENDAR EFFORT",
                "MEMBER TYPE",
            ]
        else:
            headers = [
                "NO",
                "ACCOUNT",
                "MAIL",
                "PROJECT CODE",
                "AI PROJECT",
                "REVENUE",
                "CALENDAR EFFORT",
                "MEMBER TYPE",
            ]

        # Header row với style đẹp hơn
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thick_border

        data_start_row = 2

        # Thêm data rows với alternating colors
        for row_idx, (_, row) in enumerate(df_input.iterrows(), start=data_start_row):
            current_col = 1

            # NO
            cell = ws.cell(row=row_idx, column=current_col)
            cell.value = row_idx - 1
            cell.alignment = center_align
            cell.border = thin_border
            current_col += 1

            # MONTH (nếu có)
            if has_month_label:
                cell = ws.cell(row=row_idx, column=current_col)
                cell.value = row.get("Month_Label", "")
                cell.alignment = center_align
                cell.border = thin_border
                cell.font = Font(bold=True, size=10)
                month_col = current_col
                current_col += 1

            # ACCOUNT
            cell = ws.cell(row=row_idx, column=current_col)
            cell.value = row.get("Username", "")
            cell.alignment = left_align
            cell.border = thin_border
            current_col += 1

            # MAIL
            cell = ws.cell(row=row_idx, column=current_col)
            cell.value = row.get("MAIL", "")
            cell.alignment = left_align
            cell.border = thin_border
            current_col += 1

            # PROJECT CODE
            cell = ws.cell(row=row_idx, column=current_col)
            cell.value = row.get("Project Code", "")
            cell.alignment = left_align
            cell.border = thin_border
            current_col += 1

            # AI PROJECT
            cell = ws.cell(row=row_idx, column=current_col)
            ai_value = row.get("AI Project", "")
            cell.value = ai_value
            cell.alignment = center_align
            cell.border = thin_border

            # Highlight AI projects
            if ai_value == "AI":
                cell.fill = PatternFill(
                    start_color=COLORS["ai_project"],
                    end_color=COLORS["ai_project"],
                    fill_type="solid",
                )
                cell.font = Font(bold=True, color="FF6B35")

            ai_col = current_col
            current_col += 1

            # REVENUE
            cell = ws.cell(row=row_idx, column=current_col)
            project_code = row.get("Project Code", "")
            revenue_formula = self.get_revenue_formula(project_code)
            if isinstance(revenue_formula, str) and revenue_formula.startswith("="):
                cell.value = revenue_formula
            else:
                cell.value = revenue_formula
            cell.number_format = INTEGER_FORMAT
            cell.alignment = right_align
            cell.border = thin_border
            revenue_col = current_col
            current_col += 1

            # CALENDAR EFFORT
            cell = ws.cell(row=row_idx, column=current_col)
            cell.value = row.get("Calendar Effort", 0)
            cell.number_format = NUMBER_FORMAT
            cell.alignment = right_align
            cell.border = thin_border
            effort_col = current_col
            current_col += 1

            # MEMBER TYPE
            member_type = row.get("Member Type", "Internal")
            cell = ws.cell(row=row_idx, column=current_col)
            cell.value = member_type
            cell.alignment = center_align
            cell.border = thin_border
            member_col = current_col

            # Row coloring based on member type
            color = COLORS["internal"] if member_type == "Internal" else COLORS["xjobs"]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # Tô màu các cột (trừ AI PROJECT và REVENUE)
            for col in range(1, current_col + 1):
                if col not in [ai_col, revenue_col]:
                    ws.cell(row=row_idx, column=col).fill = fill

        data_end_row = data_start_row + len(df_input) - 1

        # === TẠO BẢNG TOTAL SUMMARY ĐẸP ===
        summary_start_col = (
            11 if has_month_label else 10
        )  # Dịch sang phải nếu có cột MONTH
        summary_start_row = 1

        # Title header cho bảng summary
        title_fill = PatternFill(
            start_color="1F4E78",  # Navy blue
            end_color="1F4E78",
            fill_type="solid",
        )
        title_font = Font(bold=True, size=14, color="FFFFFF")

        ws.cell(row=summary_start_row, column=summary_start_col).value = (
            "📊 TOTAL SUMMARY"
        )
        ws.cell(row=summary_start_row, column=summary_start_col).fill = title_fill
        ws.cell(row=summary_start_row, column=summary_start_col).font = title_font
        ws.cell(row=summary_start_row, column=summary_start_col).alignment = (
            center_align
        )
        ws.cell(row=summary_start_row, column=summary_start_col).border = thick_border
        ws.merge_cells(
            start_row=summary_start_row,
            start_column=summary_start_col,
            end_row=summary_start_row,
            end_column=summary_start_col + 1,
        )

        summary_row = summary_start_row + 1

        # Style cho summary rows
        label_fill = PatternFill(
            start_color="D9E1F2",  # Light blue
            end_color="D9E1F2",
            fill_type="solid",
        )
        label_font = Font(bold=True, size=11, color="1F4E78")

        value_fill = PatternFill(
            start_color="FFFFFF",
            end_color="FFFFFF",
            fill_type="solid",
        )
        value_font = Font(bold=True, size=11, color="000000")

        # Xác định column letters dựa trên có MONTH hay không
        revenue_col_letter = get_column_letter(7 if has_month_label else 6)
        ai_col_letter = get_column_letter(6 if has_month_label else 5)
        effort_col_letter = get_column_letter(8 if has_month_label else 7)
        member_col_letter = get_column_letter(9 if has_month_label else 8)

        # Total Revenue
        ws.cell(row=summary_row, column=summary_start_col).value = "💰 Total Revenue"
        ws.cell(row=summary_row, column=summary_start_col).fill = label_fill
        ws.cell(row=summary_row, column=summary_start_col).font = label_font
        ws.cell(row=summary_row, column=summary_start_col).alignment = left_align
        ws.cell(row=summary_row, column=summary_start_col).border = thin_border

        cell = ws.cell(row=summary_row, column=summary_start_col + 1)
        cell.value = f"=SUM({revenue_col_letter}{data_start_row}:{revenue_col_letter}{data_end_row})"
        cell.number_format = "#,##0"
        cell.fill = value_fill
        cell.font = value_font
        cell.alignment = right_align
        cell.border = thin_border

        summary_row += 1

        # Total AI Revenue
        ws.cell(row=summary_row, column=summary_start_col).value = "🤖 Total AI Revenue"
        ws.cell(row=summary_row, column=summary_start_col).fill = label_fill
        ws.cell(row=summary_row, column=summary_start_col).font = label_font
        ws.cell(row=summary_row, column=summary_start_col).alignment = left_align
        ws.cell(row=summary_row, column=summary_start_col).border = thin_border

        cell = ws.cell(row=summary_row, column=summary_start_col + 1)
        cell.value = f'=SUMIF({ai_col_letter}{data_start_row}:{ai_col_letter}{data_end_row},"AI",{revenue_col_letter}{data_start_row}:{revenue_col_letter}{data_end_row})'
        cell.number_format = "#,##0"
        cell.fill = value_fill
        cell.font = value_font
        cell.alignment = right_align
        cell.border = thin_border

        summary_row += 1

        # Total Effort
        ws.cell(row=summary_row, column=summary_start_col).value = "⏱️ Total Effort"
        ws.cell(row=summary_row, column=summary_start_col).fill = label_fill
        ws.cell(row=summary_row, column=summary_start_col).font = label_font
        ws.cell(row=summary_row, column=summary_start_col).alignment = left_align
        ws.cell(row=summary_row, column=summary_start_col).border = thin_border

        cell = ws.cell(row=summary_row, column=summary_start_col + 1)
        cell.value = f"=SUM({effort_col_letter}{data_start_row}:{effort_col_letter}{data_end_row})"
        cell.number_format = NUMBER_FORMAT
        cell.fill = value_fill
        cell.font = value_font
        cell.alignment = right_align
        cell.border = thin_border

        summary_row += 1

        # Separator line
        summary_row += 1

        # Total Internal Member
        ws.cell(row=summary_row, column=summary_start_col).value = "👥 Internal Members"
        ws.cell(row=summary_row, column=summary_start_col).fill = label_fill
        ws.cell(row=summary_row, column=summary_start_col).font = label_font
        ws.cell(row=summary_row, column=summary_start_col).alignment = left_align
        ws.cell(row=summary_row, column=summary_start_col).border = thin_border

        cell = ws.cell(row=summary_row, column=summary_start_col + 1)
        cell.value = f'=COUNTIF({member_col_letter}{data_start_row}:{member_col_letter}{data_end_row},"Internal")'
        cell.fill = value_fill
        cell.font = value_font
        cell.alignment = right_align
        cell.border = thin_border

        summary_row += 1

        # Total X-Jobs Member
        ws.cell(row=summary_row, column=summary_start_col).value = "🔧 X-Jobs Members"
        ws.cell(row=summary_row, column=summary_start_col).fill = label_fill
        ws.cell(row=summary_row, column=summary_start_col).font = label_font
        ws.cell(row=summary_row, column=summary_start_col).alignment = left_align
        ws.cell(row=summary_row, column=summary_start_col).border = thin_border

        cell = ws.cell(row=summary_row, column=summary_start_col + 1)
        cell.value = f'=COUNTIF({member_col_letter}{data_start_row}:{member_col_letter}{data_end_row},"X-Jobs")'
        cell.fill = value_fill
        cell.font = value_font
        cell.alignment = right_align
        cell.border = thin_border

        summary_row += 1

        # Total Members
        ws.cell(row=summary_row, column=summary_start_col).value = "📈 Total Members"
        ws.cell(row=summary_row, column=summary_start_col).fill = PatternFill(
            start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"
        )
        ws.cell(row=summary_row, column=summary_start_col).font = Font(
            bold=True, size=11, color="1F4E78"
        )
        ws.cell(row=summary_row, column=summary_start_col).alignment = left_align
        ws.cell(row=summary_row, column=summary_start_col).border = thick_border

        cell = ws.cell(row=summary_row, column=summary_start_col + 1)
        cell.value = f"=COUNTA({member_col_letter}{data_start_row}:{member_col_letter}{data_end_row})"
        cell.fill = PatternFill(
            start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"
        )
        cell.font = Font(bold=True, size=11, color="1F4E78")
        cell.alignment = right_align
        cell.border = thick_border

        # Set column widths
        ws.column_dimensions["A"].width = 6

        if has_month_label:
            ws.column_dimensions["B"].width = 15  # MONTH
            ws.column_dimensions["C"].width = 18  # ACCOUNT
            ws.column_dimensions["D"].width = 28  # MAIL
            ws.column_dimensions["E"].width = 25  # PROJECT CODE
            ws.column_dimensions["F"].width = 13  # AI PROJECT
            ws.column_dimensions["G"].width = 15  # REVENUE
            ws.column_dimensions["H"].width = 16  # CALENDAR EFFORT
            ws.column_dimensions["I"].width = 14  # MEMBER TYPE
            ws.column_dimensions["J"].width = 3  # Spacing
        else:
            ws.column_dimensions["B"].width = 18  # ACCOUNT
            ws.column_dimensions["C"].width = 28  # MAIL
            ws.column_dimensions["D"].width = 25  # PROJECT CODE
            ws.column_dimensions["E"].width = 13  # AI PROJECT
            ws.column_dimensions["F"].width = 15  # REVENUE
            ws.column_dimensions["G"].width = 16  # CALENDAR EFFORT
            ws.column_dimensions["H"].width = 14  # MEMBER TYPE
            ws.column_dimensions["I"].width = 3  # Spacing

        ws.column_dimensions[get_column_letter(summary_start_col)].width = 22
        ws.column_dimensions[get_column_letter(summary_start_col + 1)].width = 18

        # Freeze panes
        ws.freeze_panes = "A2"

    def _create_summary_sheet(self, df_monthly, month_list):
        """Tạo sheet Summary với metrics theo tháng dùng Excel formulas"""

        if self.workbook is None:
            return

        ws = self.workbook.create_sheet(title="Summary", index=2)

        header_fill = PatternFill(
            start_color=COLORS["fixed_header"],
            end_color=COLORS["fixed_header"],
            fill_type="solid",
        )
        header_font = Font(bold=True, size=11, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")

        section_fill = PatternFill(
            start_color=COLORS["header_month"],
            end_color=COLORS["header_month"],
            fill_type="solid",
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        current_row = 1

        # Header row
        headers = ["Metrics"] + [f"{self.get_month_name(m)} {y}" for y, m in month_list]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        current_row += 1

        # Lưu vị trí các row cho việc tính toán
        total_revenue_row = current_row
        ai_revenue_row = current_row + 1
        actual_member_row = current_row + 2
        actual_member_ai_row = current_row + 3
        productivity_row = current_row + 4
        productivity_ai_row = current_row + 5
        xjob_member_row = current_row + 6
        bmm_row = current_row + 7

        # === TOTAL REVENUE ===
        ws.cell(row=current_row, column=1).value = "Total Revenue"
        ws.cell(row=current_row, column=1).fill = section_fill
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).border = thin_border

        for col_idx, (year, month) in enumerate(month_list, start=2):
            cell = ws.cell(row=current_row, column=col_idx)
            mask = (df_monthly["Year"] == year) & (df_monthly["Month"] == month)
            month_data = df_monthly[mask]

            if not month_data.empty:
                revenue = month_data["REVxEFF"].sum()
                cell.value = revenue
                cell.number_format = NUMBER_FORMAT
                cell.border = thin_border

        current_row += 1

        # === AI REVENUE ===
        ws.cell(row=current_row, column=1).value = "AI Revenue"
        ws.cell(row=current_row, column=1).fill = section_fill
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).border = thin_border

        for col_idx, (year, month) in enumerate(month_list, start=2):
            mask = (
                (df_monthly["Year"] == year)
                & (df_monthly["Month"] == month)
                & (df_monthly["AI Project"] == "AI")
            )
            month_data = df_monthly[mask]

            if not month_data.empty:
                ai_revenue = month_data["REVxEFF"].sum()
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = ai_revenue
                cell.number_format = NUMBER_FORMAT
                cell.border = thin_border

        current_row += 1

        # === ACTUAL MEMBER ===
        ws.cell(row=current_row, column=1).value = "Actual Member"
        ws.cell(row=current_row, column=1).fill = section_fill
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).border = thin_border

        for col_idx, (year, month) in enumerate(month_list, start=2):
            mask = (df_monthly["Year"] == year) & (df_monthly["Month"] == month)
            month_data = df_monthly[mask]
            count = month_data["Username"].nunique()

            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = count
            cell.border = thin_border

        current_row += 1

        # === ACTUAL MEMBER (AI) ===
        ws.cell(row=current_row, column=1).value = "Actual Member (AI)"
        ws.cell(row=current_row, column=1).fill = section_fill
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).border = thin_border

        for col_idx, (year, month) in enumerate(month_list, start=2):
            mask = (
                (df_monthly["Year"] == year)
                & (df_monthly["Month"] == month)
                & (df_monthly["AI Project"] == "AI")
            )
            month_data = df_monthly[mask]
            count = month_data["Username"].nunique()

            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = count
            cell.border = thin_border

        current_row += 1

        # === PRODUCTIVITY (dùng formula tham chiếu đến các row trên) ===
        ws.cell(row=current_row, column=1).value = "Productivity"
        ws.cell(row=current_row, column=1).fill = section_fill
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).border = thin_border

        for col_idx in range(2, len(month_list) + 2):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=current_row, column=col_idx)
            # Productivity = Total Revenue / Actual Member
            cell.value = f"=IF({col_letter}{actual_member_row}=0,0,{col_letter}{total_revenue_row}/{col_letter}{actual_member_row})"
            cell.number_format = NUMBER_FORMAT
            cell.border = thin_border

        current_row += 1

        # === PRODUCTIVITY (AI) (dùng formula) ===
        ws.cell(row=current_row, column=1).value = "Productivity (AI)"
        ws.cell(row=current_row, column=1).fill = section_fill
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).border = thin_border

        for col_idx in range(2, len(month_list) + 2):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=current_row, column=col_idx)
            # Productivity AI = AI Revenue / Actual Member (AI)
            cell.value = f"=IF({col_letter}{actual_member_ai_row}=0,0,{col_letter}{ai_revenue_row}/{col_letter}{actual_member_ai_row})"
            cell.number_format = NUMBER_FORMAT
            cell.border = thin_border

        current_row += 1

        # === X-JOB MEMBER ===
        ws.cell(row=current_row, column=1).value = "X-Job Member"
        ws.cell(row=current_row, column=1).fill = section_fill
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).border = thin_border

        for col_idx, (year, month) in enumerate(month_list, start=2):
            mask = (
                (df_monthly["Year"] == year)
                & (df_monthly["Month"] == month)
                & (df_monthly["Member Type"] == "X-Jobs")
            )
            month_data = df_monthly[mask]
            count = month_data["Username"].nunique()

            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = count
            cell.border = thin_border

        current_row += 1

        # === BMM (dùng formula tham chiếu) ===
        ws.cell(row=current_row, column=1).value = "BMM"
        ws.cell(row=current_row, column=1).fill = section_fill
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).border = thin_border

        for col_idx in range(2, len(month_list) + 2):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=current_row, column=col_idx)
            # BMM = Actual Member (giống nhau vì đã count unique)
            cell.value = f"={col_letter}{actual_member_row}"
            cell.border = thin_border

        # Set column widths
        ws.column_dimensions["A"].width = 25
        for col_idx in range(2, len(month_list) + 2):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15
